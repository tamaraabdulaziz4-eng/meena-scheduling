#!/usr/bin/env python3
"""Fill DONE / NOT DONE with the strict "same visit" rule, one branch, no cancelled rows.

Rule per order line (the reviewer's 100%-sure rule):
    DONE      = the same patient has a PACS study WITH images, of the same
                modality family, in the same visit window as the order
                (from 2 h before the order time up to 24 h after it).
                A sibling line of the same order imaged as one study
                (e.g. ankle + foot) therefore counts as DONE too.
    NOT DONE  = nothing with images for that patient/modality in that window.
Cancelled orders are dropped from the output.

Output: <export>_<Branch>.xlsx with the original columns (+ DONE OR NOT DONE)
and the branch's non-cancelled rows that are visible under the export's Excel
filter, plus a small "Why" column with the
PACS evidence (accession, exam, time, images).

Usage:
    python3 sameday_fill.py N3_JUL.xlsx N3_AUG.xlsx --pacs pacs_jul_aug.json pacs_jul_aug_part2.json \
        --branch Rawdah --out filled
"""
from __future__ import annotations

import argparse
import collections
import datetime as dt
import os
import re
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from pacs_match import DONE, NOT_DONE, VERDICT_COL, family_for, load_export, load_pacs, visible_rows  # noqa: E402

BEFORE_H, AFTER_H = 2, 24
FILL_DONE = PatternFill("solid", fgColor="C6EFCE")
FILL_NOT = PatternFill("solid", fgColor="FFC7CE")


def to_dt(v) -> dt.datetime | None:
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v
    s = str(v).strip().replace("T", " ")
    for fmt, n in (("%Y-%m-%d %H:%M:%S", 19), ("%Y-%m-%d %H:%M", 16), ("%Y-%m-%d", 10)):
        try:
            return dt.datetime.strptime(s[:n], fmt)
        except ValueError:
            continue
    return None


def decide(o: dict, studies: list[dict]) -> tuple[str, dict | None, str]:
    """Return (verdict, study used, reason).

    verdict is "" when the two reviewer rules disagree and a human must decide.
    """
    od = to_dt(o.get("Order Date"))
    if od is None:
        return NOT_DONE, None, "order has no date"
    fam = family_for(o.get("Modality (Category)"))
    code = str(o.get("Service Code") or "").strip()
    lo, hi = od - dt.timedelta(hours=BEFORE_H), od + dt.timedelta(hours=AFTER_H)
    window = []
    for s in studies:
        sd = to_dt(s["start_full"])
        if sd is None or not (lo <= sd <= hi):
            continue
        if fam is not None and s["mod"] not in fam:
            continue
        window.append((sd, s))
    window.sort(key=lambda x: x[0])
    with_img = [s for _sd, s in window if s["img"] > 0]
    own_all = [s for _sd, s in window if code and s["code"] == code]
    own_img = [s for s in own_all if s["img"] > 0]
    if own_img:
        s = own_img[0]
        return DONE, s, f"CONFIRMED DONE: own exam {s['status']} with {s['img']} images"
    own_zero = own_all[0] if own_all else None
    if with_img:
        s = max(with_img, key=lambda x: x["img"])
        if own_zero is not None:
            # rule conflict: the ordered exam itself is registered with 0 images,
            # but another exam of the same visit has images -> human decides
            return "", s, (f"CHECK: own exam registered {own_zero['status']} with 0 images "
                           f"({own_zero['acc'] or 'no accession'}), but images exist under {s['text']} "
                           f"({s['img']} images) in the same visit")
        return DONE, s, f"DONE: no separate entry for this exam; images in same visit under {s['text']} ({s['status']}, {s['img']} images)"
    if own_zero is not None:
        if own_zero["status"] == "Ordered":
            return NOT_DONE, own_zero, "CONFIRMED NOT DONE: same exam registered same day as Ordered with 0 images"
        return "", own_zero, f"CHECK: same exam registered same day as {own_zero['status']} with 0 images (arrived but no images sent?)"
    zero = [s for _sd, s in window]
    if zero:
        return NOT_DONE, zero[0], f"CONFIRMED NOT DONE: no images at all in this visit (entries: {', '.join(sorted(set(x['status'] for x in zero)))})"
    return NOT_DONE, None, "CONFIRMED NOT DONE: nothing in PACS for this visit"


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("exports", nargs="+")
    ap.add_argument("--pacs", nargs="+", required=True)
    ap.add_argument("--branch", required=True, help="text contained in the Branch column, e.g. Rawdah")
    ap.add_argument("--out", default="filled")
    ap.add_argument("--keep-cancelled", action="store_true")
    ap.add_argument("--all-rows", action="store_true", help="ignore the Excel filter of the export (default: only visible rows)")
    args = ap.parse_args(argv)
    os.makedirs(args.out, exist_ok=True)

    studies = load_pacs(args.pacs)
    # load_pacs keeps only the first 16 chars of the start time; rebuild the full stamp
    by_mrn = collections.defaultdict(list)
    for s in studies:
        s["start_full"] = s["start"]
        by_mrn[s["mrn"]].append(s)

    for path in args.exports:
        header, rows = load_export(path)
        if not args.all_rows:
            keep = visible_rows(path)
            rows = [o for i, o in enumerate(rows) if i in keep]
        rows = [o for o in rows if args.branch.lower() in str(o.get("Branch") or "").lower()]
        if not args.keep_cancelled:
            rows = [o for o in rows if not str(o.get("Order Status") or "").lower().startswith("cancel")]
        cols = list(header)
        if VERDICT_COL not in cols:
            cols.append(VERDICT_COL)
        cols.append("Why")
        cols.append("Review")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Radiology Orders"
        ws.append(cols)
        vcol = cols.index(VERDICT_COL) + 1
        counts = collections.Counter()
        reasons = collections.Counter()
        for o in rows:
            v, s, why = decide(o, by_mrn.get(str(o.get("MRNO") or "").strip(), []))
            counts[v] += 1
            reasons[why.split(":")[0] if ":" in why else why.split(" (")[0]] += 1
            evid = f"{s['acc'] or '(no accession)'} | {s['text']} | {s['start']} | {s['img']} images" if s else ""
            row = []
            for c in cols:
                if c == VERDICT_COL:
                    row.append(v)
                elif c == "Why":
                    row.append(f"{why}; {evid}" if evid else why)
                elif c == "Review":
                    row.append("CHECK" if why.startswith("CHECK") else "SURE")
                else:
                    row.append(o.get(c))
            ws.append(row)
            ws.cell(row=ws.max_row, column=vcol).fill = FILL_DONE if v == DONE else (FILL_NOT if v == NOT_DONE else PatternFill("solid", fgColor="FFEB9C"))
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 16 if c not in ("Patient Name", "Exam / Service", "Why", "Branch") else 40
        base = os.path.splitext(os.path.basename(path))[0]
        out = os.path.join(args.out, f"{base}_{re.sub(r'[^A-Za-z0-9]+', '', args.branch.title())}.xlsx")
        wb.save(out)
        # separate small workbook with only the rows a human must decide
        wc = openpyxl.Workbook(); wsc = wc.active; wsc.title = "For your review"
        wsc.append(cols)
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[cols.index("Review")] == "CHECK":
                wsc.append(list(r))
        for c in wsc[1]:
            c.font = Font(bold=True)
        wsc.freeze_panes = "A2"; wsc.auto_filter.ref = wsc.dimensions
        for i, c in enumerate(cols, start=1):
            wsc.column_dimensions[get_column_letter(i)].width = 16 if c not in ("Patient Name", "Exam / Service", "Why", "Branch") else 45
        wc.save(out.replace(".xlsx", "_FOR_REVIEW.xlsx"))
        print(f"{out}: {len(rows)} rows, DONE {counts[DONE]}, NOT DONE {counts[NOT_DONE]}, CHECK {counts['']}")
        for k, v in reasons.most_common():
            print(f"    {v:5d}  {k}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
