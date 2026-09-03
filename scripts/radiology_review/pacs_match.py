#!/usr/bin/env python3
"""Auto-fill "DONE OR NOT DONE" in radiology order exports using a PACS worklist dump.

Inputs
------
* One or more order exports (N3_JUL.xlsx, N3_AUG.xlsx ...) with the columns
  Order Date, Order ID, MRNO, Branch, Exam / Service, Service Code,
  Modality (Category), Order Status ...
* One or more PACS worklist JSON dumps captured from the GE Universal Viewer
  browser console (the ``pacs_jul_aug.json`` / ``..._part2.json`` files). Each
  dump is ``{"pages": [{"data": [[{workItem: {...}}, ...], {...count...}]}]}``.
  A flat list of already-extracted rows is accepted too.

Method
------
Every PACS study (requested procedure) is assigned to at most ONE order:

1. exact match: same MRN + same procedure/service code, study date within
   -3 .. +45 days of the order date (same day preferred, then closest);
2. modality match: same MRN + compatible modality family, study date within
   -1 .. +45 days (used when PACS registered the exam under a generic code
   such as "US" or "DX CHEST");
3. leftover studies are only used to flag cancelled orders that still have
   images.

Verdict per order (simple rule): a matched PACS study with images -> DONE,
anything else -> NOT DONE. The Notes column says why (cancelled, registered
with 0 images, not in PACS) and flags rows where the system disagrees with
PACS (result released but no images / cancelled but images exist).

Outputs (in --out): ``<export>_REVIEWED.xlsx`` per export (original columns +
verdict + evidence, a "Needs Check" sheet and a "Summary" sheet) and
``PACS_unmatched_studies.xlsx`` (studies with images that matched no order).

Usage:
    python3 pacs_match.py N3_JUL.xlsx N3_AUG.xlsx --pacs pacs_jul_aug.json pacs_jul_aug_part2.json --out review_out

Patient data never leaves the machine; nothing here is committed to git.
"""
from __future__ import annotations

import argparse
import collections
import datetime as dt
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DONE, NOT_DONE = "DONE", "NOT DONE"
VERDICT_COL = "DONE OR NOT DONE"
REVIEW_COLS = [
    "PACS Accession", "PACS Exam", "PACS Exam Date", "PACS Images", "PACS Status", "Notes",
]

# N3 "Modality (Category)" -> compatible PACS modality codes
MODALITY_FAMILY = {
    "XR": {"DX", "CR", "DR", "RF", "XA", "OT"},
    "ULTRASOUND": {"US"},
    "CT": {"CT"},
    "MRI": {"MR"},
    "MAMM": {"MG"},
    "BMD": {"BM", "OT"},
    "FLUROSCOPY": {"RF", "XA", "DX"},
    "FLUOROSCOPY": {"RF", "XA", "DX"},
    "RADIOLOGY": None,  # any modality
}

EXACT_BEFORE, EXACT_AFTER = 3, 45      # days around the order date (exact code)
FAMILY_BEFORE, FAMILY_AFTER = 1, 45    # days around the order date (modality match)

FILL_DONE = PatternFill("solid", fgColor="C6EFCE")
FILL_NOT = PatternFill("solid", fgColor="FFC7CE")
FILL_CHECK = PatternFill("solid", fgColor="FFEB9C")
FILL_HEAD = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)


# --------------------------------------------------------------------------- load
def to_date(v) -> dt.date | None:
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()[:10]
    try:
        return dt.date.fromisoformat(s)
    except ValueError:
        return None


def load_pacs(paths: list[str]) -> list[dict]:
    """Return de-duplicated PACS studies with a valid 8-digit MRN."""
    rows: dict = {}
    for path in paths:
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        items = []
        if isinstance(data, dict) and "pages" in data:
            for page in data["pages"]:
                blocks = page.get("data", []) if isinstance(page, dict) else []
                for block in blocks:
                    if isinstance(block, list):
                        items.extend(block)
        elif isinstance(data, list):
            items = data
        for it in items:
            w = it.get("workItem", it)
            rpid = w.get("requestedProcedureId") or w.get("rpid")
            ident = w.get("patientIdentifier")
            if isinstance(ident, list):
                mrn = ident[0]["value"] if ident else ""
            else:
                mrn = w.get("mrn", "")
            mrn = str(mrn).strip()
            if not re.fullmatch(r"\d{8}", mrn):
                continue
            start = w.get("procedureStartTime", w.get("start", ""))
            name = w.get("patientName")
            rows[rpid] = {
                "rpid": rpid,
                "mrn": mrn,
                "acc": w.get("procedureAccessionNo", w.get("acc", "")),
                "code": str(w.get("procedureCode", w.get("code", "")) or "").strip(),
                "text": w.get("procedureText", w.get("text", "")),
                "status": w.get("procedureStatus", w.get("status", "")),
                "date": to_date(start),
                "start": str(start)[:16],
                "mod": w.get("procedureModality", w.get("mod", "")),
                "img": int(w.get("imageCount", w.get("img", 0)) or 0),
                "name": name.get("text", "") if isinstance(name, dict) else (name or ""),
                "order": None,  # assigned order key
            }
    return list(rows.values())


def load_export(path: str) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    rows = []
    for r in it:
        if all(v in (None, "") for v in r):
            break
        rows.append(dict(zip(header, r)))
    wb.close()
    header = [h for h in header if h]
    return header, rows


# --------------------------------------------------------------------------- match
def family_for(category) -> set | None:
    key = str(category or "").strip().upper()
    if key in MODALITY_FAMILY:
        return MODALITY_FAMILY[key]
    return None if key == "" else set()


def match(orders: list[dict], studies: list[dict]) -> None:
    """Greedy one-to-one assignment; writes o['_study'], o['_match'] in place."""
    by_mrn = collections.defaultdict(list)
    for s in studies:
        by_mrn[s["mrn"]].append(s)

    for o in orders:
        o["_mrn"] = str(o.get("MRNO") or "").strip()
        o["_code"] = str(o.get("Service Code") or "").strip()
        o["_date"] = to_date(o.get("Order Date"))
        o["_cancelled"] = str(o.get("Order Status") or "").strip().lower().startswith("cancel")
        o["_study"] = None
        o["_match"] = ""
        o["_flag"] = []

    def candidates(o, before, after, exact_code):
        if o["_date"] is None:
            return []
        fam = None if exact_code else family_for(o.get("Modality (Category)"))
        out = []
        for s in by_mrn.get(o["_mrn"], []):
            if s["order"] is not None or s["date"] is None:
                continue
            delta = (s["date"] - o["_date"]).days
            if delta < -before or delta > after:
                continue
            if exact_code:
                if s["code"] != o["_code"] or not o["_code"] or o["_code"].lower() == "none":
                    continue
            else:
                if fam is not None and s["mod"] not in fam:
                    continue
            out.append((abs(delta), 0 if s["img"] > 0 else 1, s))
        return out

    def run_pass(pool, before, after, exact_code, label):
        pairs = []
        for o in pool:
            if o["_study"] is not None:
                continue
            for absd, noimg, s in candidates(o, before, after, exact_code):
                pairs.append((absd, noimg, o["_seq"], s))
        pairs.sort(key=lambda p: (p[0], p[1], p[2]))
        seq_to_order = {o["_seq"]: o for o in pool}
        for absd, _noimg, seq, s in pairs:
            o = seq_to_order[seq]
            if o["_study"] is not None or s["order"] is not None:
                continue
            o["_study"] = s
            s["order"] = o["_key"]
            if label == "exact":
                o["_match"] = "Exact code, same day" if absd == 0 else f"Exact code, {absd} day(s) apart"
            else:
                o["_match"] = "Modality match (exam name differs)" if absd == 0 else f"Modality match, {absd} day(s) apart"

    active = [o for o in orders if not o["_cancelled"]]
    run_pass(active, EXACT_BEFORE, EXACT_AFTER, True, "exact")
    run_pass(active, FAMILY_BEFORE, FAMILY_AFTER, False, "family")
    # An order whose exact PACS entry has 0 images may still have been performed
    # under a generic exam name (e.g. "US", "DX CHEST") the same day. Swap to
    # that study when one with images is still unassigned.
    for o in active:
        s0 = o["_study"]
        if s0 is None or s0["img"] > 0:
            continue
        with_img = [c for c in candidates(o, FAMILY_BEFORE, FAMILY_AFTER, False) if c[2]["img"] > 0]
        if not with_img:
            continue
        with_img.sort(key=lambda c: c[0])
        s1 = with_img[0][2]
        s0["order"] = None
        o["_study"] = s1
        s1["order"] = o["_key"]
        o["_match"] = "Modality match (exam name differs)" if with_img[0][0] == 0 else f"Modality match, {with_img[0][0]} day(s) apart"
        o["_flag"].append(f"PACS entry {s0['acc']} has 0 images; images found under {s1['acc']} ({s1['text']})")
    # cancelled orders: exact code, same day, leftover studies only -> flag
    cancelled = [o for o in orders if o["_cancelled"]]
    run_pass(cancelled, 0, 0, True, "exact")

    for o in orders:
        o["_has_other"] = bool(by_mrn.get(o["_mrn"]))


def verdict(o: dict) -> dict:
    """Simple rule: images in PACS -> DONE, otherwise NOT DONE. Notes explain why."""
    s = o["_study"]
    status = str(o.get("Order Status") or "").strip()
    released = status.lower().startswith("completed") or status.lower().startswith("result entered")
    out = {"Notes": "",
           "PACS Accession": s["acc"] if s else "", "PACS Exam": s["text"] if s else "",
           "PACS Exam Date": s["start"] if s else "", "PACS Images": s["img"] if s else 0,
           "PACS Status": s["status"] if s else "Not in PACS"}
    notes = list(o["_flag"])
    flags = []

    if s is not None and s["img"] > 0:
        out[VERDICT_COL] = DONE
        if o["_match"].startswith("Modality"):
            notes.append("Exam name differs in PACS (matched by modality)")
        elif not o["_match"].endswith("same day"):
            notes.append(o["_match"])
        if o["_cancelled"]:
            flags.append("CHECK: cancelled in system but PACS has images")
    else:
        out[VERDICT_COL] = NOT_DONE
        if o["_cancelled"]:
            notes.append("Cancelled in system")
        elif s is not None:
            notes.append("Registered in PACS with 0 images")
        else:
            notes.append("No PACS study for this patient/exam in the period" if o["_has_other"]
                         else "Patient not in PACS in the period")
        if released:
            flags.append("CHECK: result released in system but no images in PACS")
    out["Notes"] = "; ".join(flags + notes)
    out["_check"] = bool(flags)
    return out


# --------------------------------------------------------------------------- write
def autosize(ws, widths: dict | None = None, default=14, maxw=45):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        head = ws.cell(row=1, column=col).value
        w = (widths or {}).get(head)
        if w is None:
            longest = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 200) + 1)), default=default)
            w = min(max(default, longest + 2), maxw)
        ws.column_dimensions[letter].width = w


def style_header(ws):
    for c in ws[1]:
        c.font = BOLD
        c.fill = FILL_HEAD
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_reviewed(path_out: str, header: list[str], orders: list[dict], results: list[dict], label: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Radiology Orders"
    cols = list(header)
    if VERDICT_COL not in cols:
        cols.append(VERDICT_COL)
    cols += [c for c in REVIEW_COLS if c not in cols]
    ws.append(cols)
    v_idx = cols.index(VERDICT_COL) + 1
    check_rows = []
    for o, r in zip(orders, results):
        row = []
        for c in cols:
            if c == VERDICT_COL or c in REVIEW_COLS:
                row.append(r.get(c, ""))
            else:
                v = o.get(c)
                row.append(v)
        ws.append(row)
        rr = ws.max_row
        cell = ws.cell(row=rr, column=v_idx)
        cell.font = BOLD
        cell.fill = FILL_CHECK if r["_check"] else (FILL_DONE if r[VERDICT_COL] == DONE else FILL_NOT)
        if r["_check"]:
            check_rows.append(row)
    dv = DataValidation(type="list", formula1=f'"{DONE},{NOT_DONE}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(v_idx)}2:{get_column_letter(v_idx)}{max(ws.max_row, 2)}")
    style_header(ws)
    autosize(ws, {"Patient Name": 30, "Exam / Service": 34, "Notes": 60, "PACS Exam": 28, "Branch": 26})

    # Needs check sheet: flagged rows + medium/low confidence
    wc = wb.create_sheet("Needs Check")
    wc.append(cols)
    for row in check_rows:
        wc.append(row)
    style_header(wc)
    autosize(wc, {"Patient Name": 30, "Exam / Service": 34, "Notes": 60, "PACS Exam": 28, "Branch": 26})

    # Summary sheet
    wsum = wb.create_sheet("Summary")
    wsum.append([f"{label}: DONE / NOT DONE summary"])
    wsum["A1"].font = Font(bold=True, size=13)
    wsum.append([])
    wsum.append(["Branch", "Orders", DONE, NOT_DONE, "Needs check (yellow)"])
    per = collections.defaultdict(lambda: collections.Counter())
    for o, r in zip(orders, results):
        b = str(o.get("Branch") or "(no branch)")
        per[b]["n"] += 1
        per[b][r[VERDICT_COL]] += 1
        per[b]["check"] += int(r["_check"])
    tot = collections.Counter()
    for b in sorted(per):
        c = per[b]
        wsum.append([b, c["n"], c[DONE], c[NOT_DONE], c["check"]])
        tot.update(c)
    wsum.append(["TOTAL", tot["n"], tot[DONE], tot[NOT_DONE], tot["check"]])
    wsum[wsum.max_row][0].font = BOLD
    wsum.append([])
    wsum.append(["How the verdict was decided", "Orders"])
    how = collections.Counter(f'{r[VERDICT_COL]} | {r["Notes"] or "Images found in PACS, same exam, same day"}' for r in results)
    for k, v in how.most_common():
        wsum.append([k, v])
    wsum.append([])
    wsum.append(["Legend"])
    wsum.append(["Green", "DONE - images exist in PACS for this patient and exam"])
    wsum.append(["Red", "NOT DONE - no images in PACS (cancelled, registered with 0 images, or not in PACS)"])
    wsum.append(["Yellow", "System and PACS disagree - see Notes, check in Cerner/DE if needed"])
    for c in wsum[3]:
        c.font = BOLD
        c.fill = FILL_HEAD
    autosize(wsum, {"Branch": 34, "How the verdict was decided": 34}, default=16, maxw=70)

    wi = wb.create_sheet("Instructions")
    lines = [
        "How to use this file / طريقة الاستخدام",
        "",
        "EN:",
        "1. Sheet 'Radiology Orders' is your original export with DONE OR NOT DONE filled automatically.",
        "2. Rule: images exist in PACS = DONE, otherwise NOT DONE. Yellow = system and PACS disagree; check in Cerner/DE and correct the cell (dropdown) if needed.",
        "3. Sheet 'Needs Check' lists only the yellow rows.",
        "4. Columns PACS Accession / PACS Exam / PACS Exam Date / PACS Images show the evidence used for each row.",
        "5. Sheet 'Summary' gives totals per branch to send with the file.",
        "",
        "AR:",
        "1. ورقة Radiology Orders هي ملفك الأصلي مع تعبئة عمود DONE OR NOT DONE تلقائياً.",
        "2. القاعدة: توجد صور في PACS = DONE، وغير ذلك = NOT DONE. الأصفر = تعارض بين النظام وPACS، راجعه في Cerner أو DE وعدّل الخلية من القائمة المنسدلة إذا لزم.",
        "3. ورقة Needs Check فيها فقط الصفوف الصفراء.",
        "4. أعمدة PACS Accession / PACS Exam / PACS Exam Date / PACS Images توضح الدليل المستخدم لكل صف.",
        "5. ورقة Summary فيها الإجماليات لكل فرع.",
    ]
    for ln in lines:
        wi.append([ln])
    wi["A1"].font = Font(bold=True, size=13)
    wi.column_dimensions["A"].width = 120
    wb.save(path_out)


def write_unmatched(path_out: str, studies: list[dict], months: set[str]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PACS studies not in exports"
    ws.append(["MRN", "Patient Name", "Accession", "Procedure Code", "Exam", "Modality", "Exam Date", "Images", "PACS Status"])
    n = 0
    for s in sorted(studies, key=lambda s: (s["start"], s["mrn"])):
        if s["order"] is None and s["img"] > 0 and s["date"] and s["date"].strftime("%Y-%m") in months:
            ws.append([s["mrn"], s["name"], s["acc"], s["code"], s["text"], s["mod"], s["start"], s["img"], s["status"]])
            n += 1
    style_header(ws)
    autosize(ws, {"Patient Name": 30, "Exam": 30})
    wb.save(path_out)
    return n


# --------------------------------------------------------------------------- main
def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("exports", nargs="+", help="order export .xlsx files")
    ap.add_argument("--pacs", nargs="+", required=True, help="PACS worklist JSON dump(s)")
    ap.add_argument("--out", default="review_out", help="output directory")
    ap.add_argument("--branch", default="", help="only write orders whose Branch contains this text (matching still uses all orders)")
    args = ap.parse_args(argv)

    os.makedirs(args.out, exist_ok=True)
    studies = load_pacs(args.pacs)
    print(f"PACS studies with valid MRN: {len(studies)}")

    all_orders = []
    per_file = []
    seq = 0
    for path in args.exports:
        header, rows = load_export(path)
        for o in rows:
            o["_key"] = (path, o.get("Order ID"), seq)
            o["_seq"] = seq
            seq += 1
        all_orders.extend(rows)
        per_file.append((path, header, rows))
        print(f"{os.path.basename(path)}: {len(rows)} orders")

    match(all_orders, studies)

    months = set()
    for path, header, rows in per_file:
        if args.branch:
            rows = [o for o in rows if args.branch.lower() in str(o.get("Branch") or "").lower()]
        results = [verdict(o) for o in rows]
        base = os.path.splitext(os.path.basename(path))[0]
        if args.branch:
            base += "_" + re.sub(r"[^A-Za-z0-9]+", "", args.branch.title())
        out = os.path.join(args.out, f"{base}_REVIEWED.xlsx")
        write_reviewed(out, header, rows, results, base)
        c = collections.Counter(r[VERDICT_COL] for r in results)
        chk = sum(r["_check"] for r in results)
        print(f"  -> {out}: {len(rows)} orders, DONE {c[DONE]}, NOT DONE {c[NOT_DONE]}, needs check {chk}")
        for o in rows:
            if o["_date"]:
                months.add(o["_date"].strftime("%Y-%m"))

    out = os.path.join(args.out, "PACS_unmatched_studies.xlsx")
    n = write_unmatched(out, studies, months)
    print(f"  -> {out}: {n} PACS studies with images that matched no order")
    return 0


if __name__ == "__main__":
    sys.exit(main())
