#!/usr/bin/env python3
"""
Build per-branch radiology review workbooks from the monthly N3 order exports.

Why this exists
---------------
The monthly export lists every radiology order line. Leaders must confirm, for
each *manually registered* (Pending) order, whether the exam was actually
performed, using PACS (inside the network), Cerner, or DE (outside the
network). Neither PACS nor DE can export a list, so each lookup is a manual
search by patient file number (MRN). This script minimises that work:

  * rows the system already answers are auto-resolved
    (Completed / Result Entered -> DONE, Cancelled -> NOT DONE),
  * remaining rows are grouped per branch and sorted by MRN, so one search in
    DE/PACS covers every order line of that patient,
  * a "Patient Lookup" sheet lists each patient once with all their exams,
  * rows where the same patient has the same exam more than once in the month
    are flagged, because the exam name AND date must be checked in that case,
  * DONE / NOT DONE is a dropdown, with reviewer / source / notes columns,
    so the filled workbooks can be merged back by merge_reviews.py.

Usage
-----
    python3 build_review_workbooks.py N3_JUL.xlsx N3_AUG.xlsx --out review_out

Requires: openpyxl  (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------
DONE_COL = "DONE OR NOT DONE"
REVIEW_CHOICES = ["DONE", "NOT DONE", "UNCERTAIN"]
SOURCE_CHOICES = ["PACS", "Cerner", "DE", "Other"]

AUTO_RULES = {
    # Order Status                              -> (auto value, reason)
    "Completed (Result Released)": ("DONE", "auto: result released in system"),
    "Result Entered (Pending Authorization)": ("DONE", "auto: result entered in system"),
    "Cancelled": ("NOT DONE", "auto: cancelled in system"),
}

KEEP_COLS = [
    "Order Date", "Order ID", "Visit No", "MRNO", "Patient Name", "Phone",
    "Branch", "Exam / Service", "Service Code", "Modality (Category)",
    "Order Status",
]
FLAG_COLS = ["Patient orders this month", "Same exam repeated"]
REVIEW_COLS = [DONE_COL, "Verified In", "Exam Date Found", "Reviewer", "Notes"]
KEY_COL = "Review Key"  # hidden, used by merge_reviews.py

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONTH_ABBR = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
              "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


# ----------------------------------------------------------------------------
# Reading
# ----------------------------------------------------------------------------
def parse_dt(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def read_export(path: Path) -> tuple[list[str], list[dict]]:
    """Return (header, rows-as-dicts) from the first sheet containing 'Order Date'."""
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        header = None
        rows = []
        for raw in it:
            if header is None:
                if raw and "Order Date" in raw:
                    header = [str(c).strip() if c is not None else "" for c in raw]
                continue
            if not any(v is not None and str(v).strip() != "" for v in raw):
                continue
            rows.append({header[i]: raw[i] for i in range(min(len(header), len(raw)))})
        if header is not None:
            return header, rows
    raise SystemExit(f"{path}: no sheet with an 'Order Date' header found")


def month_label(rows: list[dict]) -> str:
    months = Counter()
    for r in rows:
        dt = parse_dt(r.get("Order Date"))
        if dt:
            months[(dt.year, dt.month)] += 1
    (y, m), _ = months.most_common(1)[0]
    return f"{MONTH_ABBR[m - 1]}_{y}"


def safe_name(s: str) -> str:
    s = re.sub(r"[^\w\- ]+", " ", str(s), flags=re.UNICODE)
    s = re.sub(r"\s+", "_", s.strip())
    return s.strip("_") or "UNKNOWN"


def make_keys(rows: list[dict]) -> None:
    """Attach a stable Review Key to each row: OrderID|ServiceCode|OrderDate|n."""
    seen = Counter()
    for r in rows:
        base = f"{r.get('Order ID')}|{r.get('Service Code')}|{r.get('Order Date')}"
        seen[base] += 1
        r[KEY_COL] = f"{base}|{seen[base]}"


# ----------------------------------------------------------------------------
# Classification
# ----------------------------------------------------------------------------
def classify(rows: list[dict]) -> None:
    """Add auto status + flags to every row (in place)."""
    pending = [r for r in rows if r.get("Order Status") not in AUTO_RULES]
    per_mrn = Counter(r.get("MRNO") for r in pending)
    per_mrn_exam = Counter((r.get("MRNO"), r.get("Service Code")) for r in pending)
    for r in rows:
        status = r.get("Order Status")
        if status in AUTO_RULES:
            r["_auto"], r["_auto_reason"] = AUTO_RULES[status]
        else:
            r["_auto"], r["_auto_reason"] = None, None
        r["Patient orders this month"] = per_mrn.get(r.get("MRNO"), 0)
        r["Same exam repeated"] = (
            "YES" if per_mrn_exam.get((r.get("MRNO"), r.get("Service Code")), 0) > 1 else ""
        )


# ----------------------------------------------------------------------------
# Writing helpers
# ----------------------------------------------------------------------------
def style_header(ws, ncols: int, row: int = 1) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def set_widths(ws, widths: dict[str, int], header: list[str]) -> None:
    for i, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)


WIDTHS = {
    "Order Date": 17, "Order ID": 10, "Visit No": 10, "MRNO": 12,
    "Patient Name": 34, "Phone": 13, "Branch": 26, "Exam / Service": 38,
    "Service Code": 12, "Modality (Category)": 12, "Order Status": 16,
    "Patient orders this month": 12, "Same exam repeated": 11,
    DONE_COL: 16, "Verified In": 12, "Exam Date Found": 15, "Reviewer": 16,
    "Notes": 30, KEY_COL: 4,
}


def write_review_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    """Rows needing manual review, sorted by MRN then date."""
    ws = wb.create_sheet(title)
    header = KEEP_COLS + FLAG_COLS + REVIEW_COLS + [KEY_COL]
    ws.append(header)
    style_header(ws, len(header))

    rows = sorted(rows, key=lambda r: (str(r.get("MRNO") or ""), str(r.get("Order Date") or "")))
    for r in rows:
        out = []
        for col in KEEP_COLS + FLAG_COLS:
            v = r.get(col)
            if col == "Order Date":
                v = parse_dt(v) or v
            out.append(v)
        out += ["", "", "", "", "", r[KEY_COL]]
        ws.append(out)

    n = len(rows)
    last = n + 1
    if n:
        col_idx = {name: i + 1 for i, name in enumerate(header)}
        L = lambda name: get_column_letter(col_idx[name])  # noqa: E731

        # Dropdowns
        dv_status = DataValidation(type="list", formula1=f'"{",".join(REVIEW_CHOICES)}"',
                                   allow_blank=True, showErrorMessage=True,
                                   errorTitle="Invalid", error="Choose DONE, NOT DONE or UNCERTAIN")
        dv_src = DataValidation(type="list", formula1=f'"{",".join(SOURCE_CHOICES)}"', allow_blank=True)
        ws.add_data_validation(dv_status)
        ws.add_data_validation(dv_src)
        dv_status.add(f"{L(DONE_COL)}2:{L(DONE_COL)}{last}")
        dv_src.add(f"{L('Verified In')}2:{L('Verified In')}{last}")

        # Formatting
        for row in ws.iter_rows(min_row=2, max_row=last, max_col=len(header)):
            for cell in row:
                cell.border = BORDER
            row[col_idx["Order Date"] - 1].number_format = "yyyy-mm-dd hh:mm"
            row[col_idx["Exam Date Found"] - 1].number_format = "yyyy-mm-dd"
            for name in REVIEW_COLS:
                row[col_idx[name] - 1].fill = REVIEW_FILL
        rng = f"A2:{get_column_letter(len(header))}{last}"
        d = f"${L(DONE_COL)}2"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{d}="DONE"'], fill=GREEN))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{d}="NOT DONE"'], fill=RED))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{d}="UNCERTAIN"'], fill=YELLOW))
        ws.conditional_formatting.add(
            f"{L('Same exam repeated')}2:{L('Same exam repeated')}{last}",
            CellIsRule(operator="equal", formula=['"YES"'], fill=FLAG_FILL, font=Font(bold=True)),
        )
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{last}"

    set_widths(ws, WIDTHS, header)
    ws.column_dimensions[get_column_letter(len(header))].hidden = True  # Review Key
    ws.freeze_panes = "F2"


def write_lookup_sheet(wb: Workbook, title: str, rows: list[dict], review_title: str) -> None:
    """One line per patient: what to search for in DE / PACS."""
    ws = wb.create_sheet(title)
    header = ["MRNO", "Patient Name", "Phone", "# order lines", "Exams to verify",
              "Order dates", "Lines reviewed", "Status"]
    ws.append(header)
    style_header(ws, len(header))

    groups: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        groups[str(r.get("MRNO") or "")].append(r)

    for i, (mrn, grp) in enumerate(sorted(groups.items()), start=2):
        grp = sorted(grp, key=lambda r: str(r.get("Order Date") or ""))
        exams = "; ".join(f"{r.get('Exam / Service')} [{r.get('Service Code')}]" for r in grp)
        dates = "; ".join(
            (parse_dt(r.get("Order Date")) or datetime.min).strftime("%d-%b") for r in grp
        )
        ws.append([mrn, grp[0].get("Patient Name"), grp[0].get("Phone"), len(grp), exams, dates])
        # Live progress from the review sheet
        rt = f"'{review_title}'"
        mrn_col = get_column_letter(KEEP_COLS.index("MRNO") + 1)
        done_col = get_column_letter(len(KEEP_COLS) + len(FLAG_COLS) + 1)
        ws.cell(row=i, column=7).value = (
            f'=COUNTIFS({rt}!${mrn_col}:${mrn_col},$A{i},{rt}!${done_col}:${done_col},"<>")'
        )
        ws.cell(row=i, column=8).value = f'=IF(G{i}>=D{i},"Complete",IF(G{i}>0,"Partial","Not started"))'

    last = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last, max_col=len(header)):
        for cell in row:
            cell.border = BORDER
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
    ws.conditional_formatting.add(f"H2:H{last}", CellIsRule(operator="equal", formula=['"Complete"'], fill=GREEN))
    ws.conditional_formatting.add(f"H2:H{last}", CellIsRule(operator="equal", formula=['"Partial"'], fill=YELLOW))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{last}"
    for col, w in zip("ABCDEFGH", [12, 34, 13, 10, 70, 22, 11, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C2"


def write_auto_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    """Rows resolved automatically from the system status (for transparency)."""
    ws = wb.create_sheet(title)
    header = KEEP_COLS + ["Result Released Date", DONE_COL, "Reason", KEY_COL]
    ws.append(header)
    style_header(ws, len(header))
    for r in sorted(rows, key=lambda r: str(r.get("Order Date") or "")):
        out = [parse_dt(r.get(c)) or r.get(c) if c == "Order Date" else r.get(c) for c in KEEP_COLS]
        out += [parse_dt(r.get("Result Released Date")) or r.get("Result Released Date"),
                r["_auto"], r["_auto_reason"], r[KEY_COL]]
        ws.append(out)
    set_widths(ws, {**WIDTHS, "Result Released Date": 17, "Reason": 32}, header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(header)):
        row[0].number_format = "yyyy-mm-dd hh:mm"
        row[len(KEEP_COLS)].number_format = "yyyy-mm-dd hh:mm"
    ws.column_dimensions[get_column_letter(len(header))].hidden = True
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    ws.freeze_panes = "F2"


def write_summary_sheet(wb: Workbook, branch: str, months: list[tuple[str, int, int, int]]) -> None:
    """months: (label, review_rows, patients, auto_rows)."""
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = f"Radiology manual-registration review — {branch}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Deadline: end of shift on Saturday. Fill the DONE OR NOT DONE column in every '<MONTH> Review' sheet."
    header = ["Month", "Lines to review", "Patients to search", "Reviewed", "DONE", "NOT DONE",
              "UNCERTAIN", "Remaining", "Auto-resolved lines"]
    ws.append([])
    ws.append(header)
    style_header(ws, len(header), row=4)
    done_col = get_column_letter(len(KEEP_COLS) + len(FLAG_COLS) + 1)
    for i, (label, n_rev, n_pat, n_auto) in enumerate(months, start=5):
        rt = f"'{label} Review'"
        rng = f"{rt}!${done_col}$2:${done_col}${n_rev + 1}"
        ws.append([label, n_rev, n_pat,
                   f'=COUNTA({rng})' if n_rev else 0,
                   f'=COUNTIF({rng},"DONE")' if n_rev else 0,
                   f'=COUNTIF({rng},"NOT DONE")' if n_rev else 0,
                   f'=COUNTIF({rng},"UNCERTAIN")' if n_rev else 0,
                   f'=B{i}-D{i}', n_auto])
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=len(header)):
        for cell in row:
            cell.border = BORDER
    for col, w in zip("ABCDEFGHI", [12, 15, 17, 11, 9, 11, 12, 11, 19]):
        ws.column_dimensions[col].width = w


INSTRUCTIONS = [
    ("How to review — English", True),
    ("1. Open the '<MONTH> Review' sheet. Rows are sorted by MRNO so every order line of a patient is together.", False),
    ("2. Take the MRNO (or use the '<MONTH> Patient Lookup' sheet) and search the patient ONCE in PACS (inside the network), Cerner, or DE (outside the network).", False),
    ("3. For each order line of that patient choose DONE or NOT DONE in the 'DONE OR NOT DONE' dropdown. Use UNCERTAIN only if you truly cannot tell; add a Note.", False),
    ("4. If 'Same exam repeated' = YES, the patient has the same exam more than once in the month: match by EXAM NAME and EXAM DATE, not just by the patient.", False),
    ("5. Fill 'Verified In' (PACS / Cerner / DE), 'Exam Date Found' (date of the study you saw) and your name in 'Reviewer'.", False),
    ("6. The 'Summary' and 'Patient Lookup' sheets update automatically as you fill the review sheet.", False),
    ("7. Do not delete, reorder or rename columns — the file is merged back automatically. Filtering and sorting are fine.", False),
    ("8. Rows already resolved by the system (result released / cancelled) are listed in '<MONTH> Auto' and do NOT need review.", False),
    ("", False),
    ("طريقة المراجعة — عربي", True),
    ("١. افتح شيت '<MONTH> Review'. الصفوف مرتبة حسب رقم الملف (MRNO) بحيث تكون كل طلبات المريض متجاورة.", False),
    ("٢. خذ رقم الملف (أو استخدم شيت '<MONTH> Patient Lookup') وابحث عن المريض مرة واحدة فقط في PACS (داخل الشبكة) أو Cerner أو DE (خارج الشبكة).", False),
    ("٣. لكل طلب لهذا المريض اختر DONE أو NOT DONE من القائمة المنسدلة في عمود 'DONE OR NOT DONE'. استخدم UNCERTAIN فقط إذا تعذر التأكد، واكتب ملاحظة.", False),
    ("٤. إذا كان عمود 'Same exam repeated' = YES فالمريض عنده نفس الفحص أكثر من مرة في الشهر: طابق باسم الفحص وتاريخ الفحص وليس بالمريض فقط.", False),
    ("٥. عبّئ 'Verified In' (PACS / Cerner / DE) و'Exam Date Found' (تاريخ الفحص الذي وجدته) واسمك في 'Reviewer'.", False),
    ("٦. شيت 'Summary' و'Patient Lookup' يتحدثان تلقائياً أثناء التعبئة.", False),
    ("٧. لا تحذف أو تعيد ترتيب أو تسمية الأعمدة — الملف يُدمج تلقائياً. الفلترة والترتيب مسموحان.", False),
    ("٨. الصفوف التي حسمها النظام (نتيجة صادرة / ملغاة) موجودة في شيت '<MONTH> Auto' ولا تحتاج مراجعة.", False),
]


def write_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions", 1)
    for i, (text, bold) in enumerate(INSTRUCTIONS, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=13 if bold else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 130


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("exports", nargs="+", type=Path, help="Monthly N3 export .xlsx files")
    ap.add_argument("--out", type=Path, default=Path("review_out"), help="Output directory")
    args = ap.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)

    # month label -> rows
    months: dict[str, list[dict]] = {}
    for path in args.exports:
        _, rows = read_export(path)
        make_keys(rows)
        classify(rows)
        label = month_label(rows)
        months[label] = rows
        n_auto = sum(1 for r in rows if r["_auto"])
        print(f"{path.name}: {len(rows)} lines, month {label}, auto-resolved {n_auto}, to review {len(rows) - n_auto}")

    month_order = sorted(months, key=lambda k: (k.split('_')[1], MONTH_ABBR.index(k.split('_')[0])))
    branches = sorted({str(r.get("Branch") or "UNKNOWN") for rows in months.values() for r in rows})

    totals = []
    for branch in branches:
        wb = Workbook()
        wb.remove(wb.active)
        summary_rows = []
        for label in month_order:
            rows = [r for r in months[label] if str(r.get("Branch") or "UNKNOWN") == branch]
            if not rows:
                continue
            review = [r for r in rows if not r["_auto"]]
            auto = [r for r in rows if r["_auto"]]
            short = label.split("_")[0]
            write_review_sheet(wb, f"{short} Review", review)
            write_lookup_sheet(wb, f"{short} Patient Lookup", review, f"{short} Review")
            if auto:
                write_auto_sheet(wb, f"{short} Auto", auto)
            n_pat = len({r.get("MRNO") for r in review})
            summary_rows.append((short, len(review), n_pat, len(auto)))
            totals.append((branch, short, len(review), n_pat, len(auto)))
        write_summary_sheet(wb, branch, summary_rows)
        write_instructions(wb)
        out = args.out / f"{safe_name(branch)}_review.xlsx"
        wb.save(out)
        print(f"  wrote {out.name}: " + ", ".join(f"{m} {n} lines/{p} patients" for m, n, p, _ in summary_rows))

    # Workload summary for the leader
    wb = Workbook()
    ws = wb.active
    ws.title = "Workload"
    ws.append(["Branch", "Month", "Lines to review", "Patients to search", "Auto-resolved lines"])
    style_header(ws, 5)
    for t in totals:
        ws.append(list(t))
    for col, w in zip("ABCDE", [32, 8, 15, 18, 19]):
        ws.column_dimensions[col].width = w
    wb.save(args.out / "_WORKLOAD_SUMMARY.xlsx")
    print(f"Done. {len(branches)} branch workbooks in {args.out}/")


if __name__ == "__main__":
    main()
